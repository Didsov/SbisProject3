"""
Экспорт результатов SQL-запросов ProjectSbis в Excel.

Назначение модуля:
- выполнить произвольный SELECT-запрос к локальной SQLite-базе;
- выполнить именованный запрос из src.sql_queries;
- передавать именованные параметры SQLite;
- сохранить результат в XLSX;
- использовать названия колонок непосредственно из SQL-запроса;
- автоматически оформить таблицу Excel;
- настроить ширину столбцов;
- включить автофильтр;
- закрепить строку заголовков.

Примеры:

    python -m src.sql_export \
        --file sql/query.sql

    python -m src.sql_export \
        --file sql/query.sql \
        --output exports/result.xlsx

    python -m src.sql_export \
        --query selection_summary \
        --param selection_id=41307

Функции:
- parse_arguments() — разобрать параметры CLI;
- parse_parameters() — преобразовать --param в словарь;
- read_sql_file() — прочитать SQL из файла;
- execute_query() — выполнить SQL и получить результат;
- export_to_excel() — создать оформленный XLSX;
- build_output_path() — определить путь экспортного файла;
- main() — точка входа CLI.
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.database import get_connection
from src.sql_queries import (
    get_available_queries,
    get_sql_query,
)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXPORT_DIRECTORY = PROJECT_ROOT / "exports"


def parse_arguments() -> argparse.Namespace:
    """
    Разобрать параметры запуска SQL-экспортера.

    Поддерживает два источника SQL:
    - --file — внешний .sql файл;
    - --query — именованный запрос из src.sql_queries.

    Дополнительно:
    - --param key=value — параметр SQLite;
    - --output — явное имя XLSX;
    - --sheet — имя листа Excel;
    - --list — показать доступные именованные запросы.

    Возвращает:
        argparse.Namespace с параметрами запуска.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Выполнить SQL-запрос к ProjectSbis "
            "и экспортировать результат в Excel."
        )
    )

    source = parser.add_mutually_exclusive_group()

    source.add_argument(
        "--file",
        type=Path,
        help="Путь к SQL-файлу.",
    )

    source.add_argument(
        "--query",
        help=(
            "Имя сохранённого запроса "
            "из src.sql_queries."
        ),
    )

    parser.add_argument(
        "--param",
        action="append",
        default=[],
        help=(
            "Параметр SQL в формате key=value. "
            "Можно передать несколько раз."
        ),
    )

    parser.add_argument(
        "--output",
        type=Path,
        help="Путь к итоговому XLSX.",
    )

    parser.add_argument(
        "--sheet",
        default="Результат",
        help="Название листа Excel.",
    )

    parser.add_argument(
        "--list",
        action="store_true",
        help="Показать сохранённые SQL-запросы.",
    )

    arguments = parser.parse_args()

    if (
        not arguments.list
        and arguments.file is None
        and arguments.query is None
    ):
        parser.error(
            "Нужно указать --file или --query"
        )

    return arguments


def parse_parameters(
    values: list[str],
) -> dict[str, object]:
    """
    Преобразовать параметры key=value в словарь SQLite.

    Числовые значения автоматически преобразуются в int,
    остальные остаются строками.

    Аргументы:
        values:
            Например:

            [
                "selection_id=41307",
                "limit=100",
            ]

    Возвращает:
        Словарь параметров для sqlite3.execute().

    Исключения:
        ValueError:
            Если параметр не содержит знак "="
            или имеет пустое имя.
    """
    parameters: dict[str, object] = {}

    for item in values:
        if "=" not in item:
            raise ValueError(
                f"Некорректный параметр: {item!r}. "
                "Ожидается key=value."
            )

        key, value = item.split(
            "=",
            maxsplit=1,
        )

        key = key.strip()
        value = value.strip()

        if not key:
            raise ValueError(
                f"Пустое имя параметра: {item!r}"
            )

        if value.isdigit():
            parameters[key] = int(value)
        else:
            parameters[key] = value

    return parameters


def read_sql_file(
    path: Path,
) -> str:
    """
    Прочитать SQL-запрос из текстового файла.

    Аргументы:
        path:
            Путь к .sql-файлу.

    Возвращает:
        Содержимое файла.

    Исключения:
        FileNotFoundError:
            Если SQL-файл отсутствует.

        ValueError:
            Если SQL-файл пуст.
    """
    sql = path.read_text(
        encoding="utf-8"
    ).strip()

    if not sql:
        raise ValueError(
            f"SQL-файл пуст: {path}"
        )

    return sql

def execute_query(
    sql: str,
    parameters: dict[str, object],
) -> tuple[list[str], list[tuple[Any, ...]]]:
    """
    Выполнить read-only SQL-запрос к локальной базе ProjectSbis.

    Что делает:
    - проверяет SQL через validate_read_only_sql();
    - открывает SQLite-базу в режиме read-only;
    - выполняет SELECT/WITH запрос;
    - получает названия колонок из cursor.description;
    - возвращает строки результата.

    Аргументы:
        sql:
            SQL-запрос.

        parameters:
            Именованные параметры SQLite.

    Возвращает:
        Кортеж:
        - список названий столбцов;
        - список строк результата.

    Исключения:
        ValueError:
            Если запрос пытается изменить базу;
            если запрос не возвращает табличный результат.
    """
    validate_read_only_sql(
        sql
    )

    from src.database import DATABASE_FILE

    database_uri = (
        DATABASE_FILE.resolve().as_uri()
        + "?mode=ro"
    )

    import sqlite3

    with sqlite3.connect(
        database_uri,
        uri=True,
    ) as connection:
        connection.row_factory = sqlite3.Row

        cursor = connection.execute(
            sql,
            parameters,
        )

        if cursor.description is None:
            raise ValueError(
                "SQL-запрос не вернул табличный результат."
            )

        columns = [
            description[0]
            for description in cursor.description
        ]

        rows = [
            tuple(row)
            for row in cursor.fetchall()
        ]

    return columns, rows

def add_information_sheet(
    workbook: Workbook,
    *,
    source_name: str,
    sql: str,
    parameters: dict[str, object],
    rows_count: int,
    output_path: Path,
) -> None:
    """
    Добавить в Excel служебный лист с информацией об экспорте.

    На лист записываются:
    - имя SQL-запроса;
    - дата и время экспорта;
    - количество строк;
    - путь итогового XLSX;
    - переданные параметры;
    - полный текст SQL-запроса.

    Аргументы:
        workbook:
            Создаваемая Excel-книга.

        source_name:
            Имя запроса или SQL-файла.

        sql:
            Полный текст выполненного SQL.

        parameters:
            Переданные параметры запроса.

        rows_count:
            Количество полученных строк.

        output_path:
            Путь итогового XLSX.

    Возвращает:
        None.
    """
    worksheet = workbook.create_sheet(
        title="Информация"
    )

    worksheet.append([
        "Параметр",
        "Значение",
    ])

    worksheet.append([
        "Запрос",
        source_name,
    ])

    worksheet.append([
        "Дата экспорта",
        datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
    ])

    worksheet.append([
        "Количество строк",
        rows_count,
    ])

    worksheet.append([
        "Файл",
        str(output_path),
    ])

    if parameters:
        parameters_text = "\n".join(
            f"{key} = {value}"
            for key, value in parameters.items()
        )
    else:
        parameters_text = "Нет"

    worksheet.append([
        "Параметры SQL",
        parameters_text,
    ])

    worksheet.append([
        "SQL",
        sql,
    ])

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True
        )

    worksheet.freeze_panes = "A2"

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 100

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

def build_output_path(
    requested_path: Path | None,
    source_name: str,
) -> Path:
    """
    Определить путь итогового XLSX.

    Если --output не указан, файл автоматически создаётся
    в каталоге exports с датой и временем.

    Аргументы:
        requested_path:
            Явно заданный путь либо None.

        source_name:
            Короткое имя SQL-запроса.

    Возвращает:
        Полный путь XLSX.
    """
    if requested_path is not None:
        path = requested_path

        if not path.is_absolute():
            path = PROJECT_ROOT / path

    else:
        timestamp = datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )

        filename = (
            f"{source_name}_{timestamp}.xlsx"
        )

        path = EXPORT_DIRECTORY / filename

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    return path


def export_to_excel(
    columns: list[str],
    rows: list[tuple[Any, ...]],
    output_path: Path,
    sheet_name: str,
    source_name: str,
    sql: str,
    parameters: dict[str, object],
) -> None:
    """
    Сохранить результат SQL в оформленный XLSX.

    Оформление:
    - первая строка содержит SQL-названия колонок;
    - заголовки выделены жирным;
    - включён автофильтр;
    - первая строка закреплена;
    - данные оформлены как Excel Table;
    - ширина каждого столбца рассчитывается автоматически;
    - длинный текст ограничивается разумной шириной;
    - заголовки и данные выравниваются по верхнему краю.

    Аргументы:
        columns:
            Названия колонок SQL-результата.

        rows:
            Строки SQL-результата.

        output_path:
            Путь создаваемого XLSX.

        sheet_name:
            Название листа Excel.

    Возвращает:
        None.
    """
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.title = sheet_name[:31]

    worksheet.append(
        columns
    )

    for row in rows:
        worksheet.append(
            list(row)
        )

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            vertical="top"
        )

    worksheet.freeze_panes = "A2"


    for row in worksheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=False,
            )

    if rows and columns:
        last_column = get_column_letter(
            len(columns)
        )

        last_row = len(rows) + 1

        table_reference = (
            f"A1:{last_column}{last_row}"
        )

        table = Table(
            displayName="SqlResult",
            ref=table_reference,
        )

        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        worksheet.add_table(
            table
        )

    for column_index, column_name in enumerate(
        columns,
        start=1,
    ):
        maximum_length = len(
            str(column_name)
        )

        for row_index in range(
            2,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_index,
                column=column_index,
            ).value

            if value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(str(value)),
            )

        width = min(
            max(maximum_length + 2, 10),
            50,
        )

        column_letter = get_column_letter(
            column_index
        )

        worksheet.column_dimensions[
            column_letter
        ].width = width

    add_information_sheet(
        workbook,
        source_name=source_name,
        sql=sql,
        parameters=parameters,
        rows_count=len(rows),
        output_path=output_path,
    )    
    workbook.save(
        output_path
    )


def main() -> None:
    """
    Запустить SQL-экспортёр из командной строки.

    Выбирает SQL из файла или реестра,
    выполняет его и создаёт XLSX.
    """
    arguments = parse_arguments()

    if arguments.list:
        queries = get_available_queries()

        if not queries:
            print(
                "Избранных SQL-запросов пока нет."
            )
            return

        print(
            "Доступные SQL-запросы:"
        )

        for query_name in queries:
            print(
                f"- {query_name}"
            )

        return

    parameters = parse_parameters(
        arguments.param
    )

    if arguments.file is not None:
        sql = read_sql_file(
            arguments.file
        )

        source_name = (
            arguments.file.stem
        )

    else:
        sql = get_sql_query(
            arguments.query
        )

        source_name = (
            arguments.query
        )

    columns, rows = execute_query(
        sql,
        parameters,
    )

    output_path = build_output_path(
        requested_path=arguments.output,
        source_name=source_name,
    )

    export_to_excel(
        columns=columns,
        rows=rows,
        output_path=output_path,
        sheet_name=arguments.sheet,
        source_name=source_name,
        sql=sql,
        parameters=parameters,
    )

    print()
    print(
        f"SQL-запрос выполнен."
    )

    print(
        f"Получено строк: {len(rows)}"
    )

    print(
        f"Excel сохранён: {output_path}"
    )


def validate_read_only_sql(
    sql: str,
) -> None:
    """
    Проверить, что SQL-запрос предназначен только для чтения.

    Разрешённые запросы:
    - SELECT;
    - WITH, если итоговый запрос является чтением данных.

    Запрещённые операции:
    - INSERT;
    - UPDATE;
    - DELETE;
    - DROP;
    - ALTER;
    - CREATE;
    - REPLACE;
    - VACUUM;
    - ATTACH;
    - DETACH;
    - PRAGMA;
    - REINDEX;
    - TRUNCATE.

    Аргументы:
        sql:
            Текст SQL-запроса.

    Возвращает:
        None.

    Исключения:
        ValueError:
            Если запрос пустой;
            если запрос не является read-only;
            если обнаружена запрещённая SQL-команда.

    Примечания:
        Проверка является дополнительным уровнем защиты.
        Основная защита обеспечивается также открытием отдельного
        SQLite-соединения в режиме read-only внутри execute_query().
    """
    normalized_sql = sql.strip()

    if not normalized_sql:
        raise ValueError(
            "SQL-запрос не может быть пустым"
        )

    # Убираем начальные SQL-комментарии, чтобы корректно определить
    # первую реальную команду запроса.
    normalized_sql = re.sub(
        r"^\s*(?:--[^\n]*\n\s*|/\*.*?\*/\s*)*",
        "",
        normalized_sql,
        flags=re.DOTALL,
    )

    first_word_match = re.match(
        r"([A-Za-z]+)",
        normalized_sql,
    )

    if first_word_match is None:
        raise ValueError(
            "Не удалось определить тип SQL-запроса"
        )

    first_word = first_word_match.group(1).upper()

    if first_word not in {
        "SELECT",
        "WITH",
    }:
        raise ValueError(
            "sql_export работает только в read-only режиме. "
            "Разрешены только SELECT и WITH-запросы."
        )

    forbidden_keywords = {
        "INSERT",
        "UPDATE",
        "DELETE",
        "DROP",
        "ALTER",
        "CREATE",
        "REPLACE",
        "VACUUM",
        "ATTACH",
        "DETACH",
        "PRAGMA",
        "REINDEX",
        "TRUNCATE",
    }

    # Убираем строковые литералы и комментарии, чтобы слово вроде
    # "DELETE" внутри текста не считалось SQL-командой.
    sql_without_comments = re.sub(
        r"--[^\n]*|/\*.*?\*/",
        " ",
        normalized_sql,
        flags=re.DOTALL,
    )

    sql_without_strings = re.sub(
        r"'(?:''|[^'])*'",
        "''",
        sql_without_comments,
    )

    words = {
        word.upper()
        for word in re.findall(
            r"\b[A-Za-z]+\b",
            sql_without_strings,
        )
    }

    detected_forbidden = sorted(
        forbidden_keywords.intersection(words)
    )

    if detected_forbidden:
        raise ValueError(
            "sql_export запрещает изменение базы данных. "
            "Обнаружены запрещённые SQL-команды: "
            + ", ".join(detected_forbidden)
        )

if __name__ == "__main__":
    main()